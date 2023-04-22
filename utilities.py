from RPA.Robocorp.WorkItems import WorkItems
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font
from config import DIRECTORIES

def get_workitems() -> dict:
    """
    Returns input workitems.
    """
    workitems = WorkItems()
    workitems.get_input_work_item()
    workitem = workitems.get_work_item_variables()
    return workitem

def create_excel(filepath) -> None:
    """
    Creates the excel file with required heads.
    :param filepath: Path to the excel file.
    """
    heads = ["Title", "Date", "Description", "Picture name", "Count of search phrase", "Contains money"]
    workbook = Workbook()
    workbook.save(filepath)
    wb = load_workbook(filepath)
    sheet: Worksheet = wb.active
    for index, head in enumerate(heads):
        sheet.cell(1, index+1).value = head
        sheet.cell(1, index+1).font = Font(bold=True)
    wb.save(filepath)

def update_excel(title: str, date: str, description: str, filename: str, title_count: int, desc_count: int, money_present: bool) -> None:
    """
    Updates the excel with news data.
    :param title: Title of the news.
    :param description: News description.
    :param filename: Name of the image file.
    :param title_count: Count of the search phrase in news title.
    :param desc_count: Count of the search phrase in news description.
    :param money_present: True if money is found in either news title or news description, else False.
    """
    workbook = load_workbook(DIRECTORIES.FILEPATH)
    sheet: Worksheet = workbook.active
    max_row = sheet.max_row
    sheet.cell(max_row+1, 1).value = title
    sheet.cell(max_row+1, 2).value = date
    sheet.cell(max_row+1, 3).value = description
    sheet.cell(max_row+1, 4).value = filename
    sheet.cell(max_row+1, 5).value = f"Title: {title_count} | Description: {desc_count}"
    sheet.cell(max_row+1, 6).value = money_present
    workbook.save(DIRECTORIES.FILEPATH)
