from RPA.Browser.Selenium import Selenium
from dateutil.relativedelta import relativedelta
from retry import retry
from time import sleep
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font
from config import FILEPATH, IMAGE_PATH
from selenium.common.exceptions import ElementClickInterceptedException
import datetime
import requests
import re

class NyTimes:
    def __init__(self, workitem: dict) -> None:
        self.browser = Selenium()
        self.phrase: str = workitem["phrase"]
        self.section: str = workitem["section"]
        self.months = workitem["months"]
        self.heads = ["Title", "Date", "Description", "Picture name", "Count of search phrase", "Contains money"]
    
    def get_date(self) -> str:
        """
        Returns a date the number of months prior to the current month.
        :param months: Number of months.
        """
        today = datetime.date.today()
        if int(self.months) == 0 or int(self.months) == 1:
            return today.strftime("%m/%d/%Y")
        three_mon_rel = relativedelta(months=int(self.months))
        req_date = (today - three_mon_rel).strftime("%m/%d/%Y")
        return req_date
    
    def create_excel(self, filepath) -> None:
        """
        Creates the excel file with required heads.
        :param filepath: Path to the excel file.
        """
        workbook = Workbook()
        workbook.save(filepath)
        wb = load_workbook(filepath)
        sheet: Worksheet = wb.active
        for index, head in enumerate(self.heads):
            sheet.cell(1, index+1).value = head
            sheet.cell(1, index+1).font = Font(bold=True)
        wb.save(filepath)

    def search_query(self):
        """
        Opens the website and searches for the query.
        """
        self.browser.open_available_browser("https://www.nytimes.com")
        self.browser.maximize_browser_window()
        self.browser.maximize_browser_window()
        self.browser.wait_until_element_is_visible("//button[@data-test-id='search-button']", 10)
        self.browser.click_element("//button[@data-test-id='search-button']")
        self.browser.input_text("//input[@name='query']", self.phrase)
        self.browser.click_element("//button[@type='submit']")
        self.browser.wait_until_element_is_visible("//ol[@data-testid='search-results']", 10)
    
    def set_dates(self) -> None:
        """
        Sets the desired dates in the date field.
        :param months: Number of months priors to the current month.
        """
        start_date = self.get_date()
        end_date = datetime.date.today().strftime("%m/%d/%Y")
        self.browser.click_element_when_visible("//label[text()='Date Range']")
        self.browser.wait_until_element_is_visible("//li[@class='css-guqk22']", 10)
        self.browser.click_element_when_visible("//button[text()='Specific Dates']")
        self.browser.wait_until_element_is_visible("//div[@class='css-79elbk']", 5)
        self.browser.input_text("//input[@data-testid='DateRange-startDate']", start_date)
        self.browser.input_text("//input[@data-testid='DateRange-endDate']", end_date)
        self.browser.press_keys("//input[@data-testid='DateRange-endDate']", "RETURN")

    def set_filters(self) -> None:
        """
        Set the desired filters on the result page.
        """
        self.set_dates()
        self.browser.click_element_when_visible("//label[text()='Section']")
        # self.browser.click_element_when_visible(f"//span[text()='{self.section}']")
        # for testing
        # sections = self.browser.get_webelements("//label[@class='css-1a8ayg6']")
        # section: str = self.browser.get_text(sections[5])
        # if "," in section:
        #     section = section.replace(",", "")
        # section = ''.join([i for i in section if not i.isdigit()])
        self.browser.click_element_when_visible(f"//span[text()='{self.section.capitalize()}']") # {self.section}
        self.browser.click_element("//label[text()='Section']")

    @retry((Exception), 5, 5)
    def load_all_news(self) -> None:
        """
        Loads all the news for the given time period.
        """
        self.browser.scroll_element_into_view("//a[@data-testid='search-result-qualtrics-link']")
        sleep(3600)
        while self.browser.is_element_visible("//button[text()='Show More']"):
            try:
                self.browser.click_element("//button[text()='Show More']")
            except ElementClickInterceptedException:
                self.browser.click_element()
            sleep(1)
            self.browser.scroll_element_into_view("//a[@data-testid='search-result-qualtrics-link']")
        all_news = self.browser.get_webelements("//h4[@class='css-2fgx4k']")
        self.browser.scroll_element_into_view(all_news[0])
    
    def download_picture(self, img_xpath: str, img_path: str) -> None:
        """
        Downlaods the image.
        :param img_xpath: Xpath of the image.
        """
        img_url = self.browser.get_element_attribute(img_xpath, "src")
        data = requests.get(img_url).content
        with open(img_path, "wb") as f:
            f.write(data)

    def get_count(self, input_string: str, query: str) -> int:
        """
        Returns count of a query in the input string.
        :param input_string: Input string.
        :param query: Query.
        """
        for char in ".,;?!‘’":
            input_string = input_string.lower().replace(char, "")
        words = input_string.split()
        result = []
        for i in range(0, len(words), len(query.split())):
            result.append(' '.join(words[i:i+len(query.split())]))
        return result.count(query.lower())

    def is_money_present(self, text: str) -> bool:
        """
        Checks if any amount of money is present in the given text.
        :param text: Input string.
        """
        pattern = r'\$\d+(?:,\d+)*(?:\.\d+)?(?:\s*(?:dollars|USD))?\b|\b\d+\s*(?:dollars|USD)\b'
        match = re.findall(pattern, text)
        if match:
            return True
        else:
            return False

    def fetch_data(self) -> None:
        """
        Fetches the data for all the news.
        """
        self.create_excel(FILEPATH)
        date_elements = self.browser.get_webelements("//span[@class='css-17ubb9w']") # list of date elements
        for index, date_element in enumerate(date_elements):
            date = self.browser.get_text(date_element)
            title = self.browser.get_text(self.browser.get_webelements("//h4[@class='css-2fgx4k']")[index])
            if self.browser.is_element_visible(f"//h4[text()='{title}']/..//p[@class='css-16nhkrn']"):
                description = self.browser.get_text(f"//h4[text()='{title}']/..//p[@class='css-16nhkrn']")
            else:
                description = "Not available"
            img_xpath = f"//div[@class='css-e1lvw9' and a/h4/text()='{title}']/following-sibling::figure[@class='css-tap2ym']/..//div/..//img[@class='css-rq4mmj']"
            if self.browser.is_element_visible(img_xpath):
                filepath = f"{IMAGE_PATH}/img_{index+1}.png"
                self.download_picture(img_xpath, filepath)
            else:
                filepath = "Not available"
            wb = load_workbook(FILEPATH)
            sheet: Worksheet = wb.active
            max_row = sheet.max_row
            sheet.cell(max_row+1, 1).value = title
            sheet.cell(max_row+1, 2).value = date
            sheet.cell(max_row+1, 3).value = description
            sheet.cell(max_row+1, 4).value = filepath
            sheet.cell(max_row+1, 5).value = f"Title: {self.get_count(title, self.phrase)} | Description: {self.get_count(description, self.phrase)}"
            sheet.cell(max_row+1, 6).value = self.is_money_present(title) or self.is_money_present(description)
            wb.save(FILEPATH)

# if __name__ == "__main__":
#     try:
#         flag = False
#         ny_times = NyTimes("Machines", "Sports", 3)
#         print(f"Searching query {ny_times.phrase}...")
#         ny_times.search_query()
#         try:
#             ny_times.set_filters()
#         except AssertionError:
#             flag = True
#             print(f"Section {ny_times.section} is not available.")
#         if not flag:
#             ny_times.load_all_news()
#             ny_times.fetch_data()
#             print("Query searched")
#         ny_times.browser.close_browser()
#     except Exception as e:
#         ny_times.browser.screenshot(filename="images/error.png")
#         raise e